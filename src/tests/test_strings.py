'''
Copyright (c) 2010 openpyxl

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
THE SOFTWARE.

@license: http://www.opensource.org/licenses/mit-license.php
@author: Eric Gazoni
'''

from tests.helper import BaseTestCase

from openpyxl.workbook import Workbook
from openpyxl.writer.strings import create_string_table

class TestWriterStrings(BaseTestCase):

    def test_create_string_table(self):

        wb = Workbook()

        ws = wb.create_sheet()

        ws.cell('B12').value = 'hello'
        ws.cell('B13').value = 'world'
        ws.cell('D28').value = 'hello'

        table = create_string_table(workbook = wb)

        self.assertEqual({'hello' : 1,
                          'world' : 0}, table)

