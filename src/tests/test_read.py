# coding=UTF-8
'''
Copyright (c) 2010 openpyxl

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
THE SOFTWARE.

@license: http://www.opensource.org/licenses/mit-license.php
@author: Eric Gazoni
'''
from __future__ import with_statement
import os.path as osp
from tests.helper import BaseTestCase, TMPDIR, DATADIR

from openpyxl.worksheet import Worksheet

from openpyxl.reader.worksheet import read_worksheet

class TestReadWorksheet(BaseTestCase):

    def test_read_worksheet(self):

        class DummyWb(object):

            def get_sheet_by_name(self, value):

                return None

        with open(osp.join(DATADIR, 'reader', 'sheet2.xml')) as f:
            content = f.read()

            ws = read_worksheet(xml_source = content,
                                parent = DummyWb(),
                                preset_title = 'Sheet 2',
                                string_table = {1 : 'hello'})

            self.assertTrue(isinstance(ws, Worksheet))

            self.assertEqual(ws.cell('G5').value, 'hello')

            self.assertEqual(ws.cell('D30').value, 30)

            self.assertEqual(ws.cell('K9').value, 0.09)
